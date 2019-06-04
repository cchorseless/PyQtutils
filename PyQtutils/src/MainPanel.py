from PyQt5 import QtWidgets, QtGui, QtCore
from core.Utils.BasePanel import BasePanel
from view.UiMainPanel import UiMainPanel
import openpyxl
from core.Utils.ExcelPanel import *
import os
from src.datalink.DataLinkPanel import DataLinkPanel
from core.Manage.SignalManage import *
import json
import time


class MainPanel(QtWidgets.QMainWindow, UiMainPanel, BasePanel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.eventListener()
        self.subPanel = None
        self.item_file_dir = None
        self.sheet_path = {}  # item.xlsx 关联的表文件路径
        self.sheet_ok = {}
        self.pushButton_2.setDisabled(True)
        self.pushButton_4.setDisabled(True)
        self.pushButton_5.setDisabled(True)
        self.pushButton_6.setDisabled(True)
        self.pushButton_7.setDisabled(True)
        self.pushButton_12.setDisabled(True)
        self.pushButton_13.setDisabled(True)
        self.pushButton_14.setDisabled(True)
        self.pushButton_15.setDisabled(True)
        self.timer = QtCore.QBasicTimer()

    def eventListener(self):
        Worker().parse_triggered.connect(self.changedata)
        self.pushButton.clicked.connect(self.openfile)
        self.pushButton_2.clicked.connect(self.loadSheet)
        self.pushButton_3.clicked.connect(self.resert)
        self.pushButton_15.clicked.connect(lambda: self.loadSheetData(self.tableWidget_4, True))
        self.pushButton_4.clicked.connect(lambda: self.loadSheetData(self.tableWidget, False))
        self.pushButton_13.clicked.connect(self.saveFile)
        self.pushButton_12.clicked.connect(self.exportClientData)
        self.pushButton_14.clicked.connect(self.exportServerData)
        self.pushButton_5.clicked.connect(lambda: self.checkDataChange(self.sheet_ok))
        self.pushButton_6.clicked.connect(self.gailvcheck)
        self.pushButton_7.clicked.connect(self.hengxiangshujucheck)

    def openfile(self):
        print(os.getcwd())
        self.item_file_dir = QtWidgets.QFileDialog.getOpenFileName(self, "item.xlsx文件路径",
                                                                   os.getcwd(),
                                                                   "Excel files(*.xlsx)")[0]
        if self.item_file_dir:
            self.label_9.setText(self.item_file_dir)
            self.pushButton_2.setEnabled(True)
            self.pushButton.setDisabled(True)

    def loadSheet(self):
        if self.item_file_dir:
            wb = openpyxl.load_workbook(self.item_file_dir, data_only=True)
            ws = wb.worksheets[0]
            self._showleftdata(ws, self.tableWidget_2)
            self._showleftdata(ws, self.tableWidget_3)
            self.pushButton_4.setEnabled(True)
            self.pushButton_15.setEnabled(True)
        else:
            warring_msg = QtWidgets.QMessageBox.information(self, "警告", "未选择item.xlsx", QtWidgets.QMessageBox.Yes)

    def _showleftdata(self, sheet, Qtable):
        max_row = len(sheet["A"])
        parent_path = os.path.dirname(self.item_file_dir)
        Qtable.setRowCount(max_row - 1)
        Qtable.setColumnCount(3)
        Qtable.setHorizontalHeaderLabels(["数据表", "文件名", "状态"])
        Qtable.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)  # 整行选中
        Qtable.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)  # 设置表格不可编辑
        Qtable.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)  # 单选模式
        self.progressBar.setRange(0, max_row - 2)
        for i in range(max_row - 1):
            self.progressBar.setValue(i)
            datadir = sheet.cell(row=i + 2, column=3).value
            sheetname = sheet.cell(row=i + 2, column=1).value
            Qtable.setItem(i, 0, QtWidgets.QTableWidgetItem(datadir))
            Qtable.setItem(i, 1, QtWidgets.QTableWidgetItem(sheetname))
            _datadir = str(datadir)
            _sheetname = str(sheetname)
            if os.path.exists(parent_path + "/" + _datadir):
                if os.path.exists(parent_path + "/" + _datadir + "/" + _sheetname + ".xlsx"):
                    Qtable.setItem(i, 2, QtWidgets.QTableWidgetItem("OK"))
                    self.sheet_path[i] = parent_path + "/" + _datadir + "/" + _sheetname + ".xlsx"
                    self.sheet_ok[i] = _datadir
                else:
                    Qtable.setItem(i, 2, QtWidgets.QTableWidgetItem("文件丢失"))
            else:
                Qtable.setItem(i, 2, QtWidgets.QTableWidgetItem("文件夹丢失"))

    def loadSheetData(self, tableWidget, bool=False):
        """
        :param tableWidget: 显示的Qtablewight
        :param bool:是否可以编辑，True 可编辑
        :return:
        """
        selectmode = 0
        selectnum = 0
        if tableWidget == self.tableWidget:
            current_row = self.tableWidget_2.currentRow()
            selectmode = 2
            selectnum = 2
        else:
            current_row = self.tableWidget_3.currentRow()
        if current_row in self.sheet_path.keys():
            self.tableWidget.clear()
            wb = openpyxl.load_workbook(self.sheet_path[current_row], data_only=True)
            ws = wb.worksheets[0]
            if ws.max_row > 20000:
                warring_msg = QtWidgets.QMessageBox.information(self, "警告", "数据表行数超过20000行，无法处理",
                                                                QtWidgets.QMessageBox.Yes)
                return
            ExceltoQTableWidgets(ws, tableWidget, bool, selectmode, selectnum, self.progressBar)
            self.pushButton_5.setEnabled(True)
            self.pushButton_6.setEnabled(True)
            self.pushButton_7.setEnabled(True)
            self.pushButton_13.setEnabled(True)
            self.pushButton_12.setEnabled(True)
            self.pushButton_14.setEnabled(True)
        else:
            warring_msg = QtWidgets.QMessageBox.information(self, "警告", "文件未选中，请检查文件状态", QtWidgets.QMessageBox.Yes)

    def checkDataChange(self, sheet_ok):
        self.subPanel = DataLinkPanel()
        self.subPanel.loaddata(sheet_ok)
        self.subPanel.show()

    def changedata(self, datalist):
        wslist = []
        for data in datalist:
            wb = openpyxl.load_workbook(self.sheet_path[data], data_only=True)
            ws = wb.worksheets[0]
            wslist.append(ws)
            if ws.max_row > 20000:
                warring_msg = QtWidgets.QMessageBox.information(self, "警告", "数据表行数超过20000行，无法处理",
                                                                QtWidgets.QMessageBox.Yes)
                return
        b = {}
        for ws in wslist:
            for i in range(ws.max_row):
                b[str(ws.cell(i + 2, 1).value)] = ws.cell(i + 2, 2).value
        maxRow = self.tableWidget.rowCount()
        curColumn = self.tableWidget.selectedIndexes()
        if len(curColumn) != maxRow:  # 判定是否多选或者没有选择
            warring_msg = QtWidgets.QMessageBox.information(self, "警告", "只能选择一列进行匹配", QtWidgets.QMessageBox.Yes)
            return
        curColumn = self.tableWidget.currentColumn()
        self.progressBar.setRange(1, maxRow - 1)
        for i in range(1, maxRow):
            self.progressBar.setValue(i)
            ErrorMark = False
            item = self.tableWidget.item(i, curColumn)
            itemtext = str(item.text())
            tmp = itemtext.split("|")
            for j in range(len(tmp)):
                if tmp[j] in b.keys():
                    tmp[j] = b[tmp[j]]
                else:
                    ErrorMark = True
            itemtext = "|".join(tmp)
            item.setText(itemtext)
            if ErrorMark:
                item.setBackground(QtGui.QColor(255, 0, 0))

    def gailvcheck(self):
        maxRow = self.tableWidget.rowCount()
        curColumn = self.tableWidget.selectedIndexes()
        if len(curColumn) != maxRow:  # 判定是否多选或者没有选择
            warring_msg = QtWidgets.QMessageBox.information(self, "警告", "只能选择一列进行概率计算",
                                                            QtWidgets.QMessageBox.Yes)
            return
        self.progressBar.setRange(1, maxRow - 1)
        curColumn = self.tableWidget.currentColumn()
        for i in range(1, maxRow):
            self.progressBar.setValue(i)
            item = self.tableWidget.item(i, curColumn)
            itemtext = str(item.text())
            tmp = itemtext.split("|")
            he = 0
            try:
                he = sum([int(x) for x in tmp])
                if he == 0:
                    raise ZeroDivisionError
            except Exception:
                item.setBackground(QtGui.QColor(255, 255, 0))
                continue
            for j in range(len(tmp)):
                tmp[j] = str(round(int(tmp[j]) / he, 2))
                itemtext = "|".join(tmp)
                item.setText(itemtext)

    def hengxiangshujucheck(self):
        maxRow = self.tableWidget.rowCount()
        curColumn = self.tableWidget.selectedIndexes()
        if len(curColumn) <= maxRow:  # 判定是否多选或者没有选择
            warring_msg = QtWidgets.QMessageBox.information(self, "警告", "只能选择多列进行匹配",
                                                            QtWidgets.QMessageBox.Yes)
            return
        curColumnlist = []  # 多选列
        for i in range(len(curColumn) // maxRow):
            curColumnlist.append(curColumn[i * maxRow].column())
        self.progressBar.setRange(1, maxRow - 1)
        for i in range(1, maxRow):
            self.progressBar.setValue(i)
            check_box = []

            for j in curColumnlist:
                item = self.tableWidget.item(i, j)
                itemtext = str(item.text())
                tmp = itemtext.split("|")
                check_box.append(len(tmp))

            if len(set(check_box)) != 1:  # 等于1说明数据量相等
                self.tableWidget.item(i, 0).setBackground(QtGui.QColor(0, 255, 0))
            else:
                self.tableWidget.item(i, 0).setBackground(QtGui.QColor(255, 255, 255))

    def saveFile(self):
        """打包成一个文件"""
        _path = os.getcwd() + '\\' + 'client'
        data = {}
        for fileName in os.listdir(_path):
            file = fileName.replace('.json', '')
            if file == 'AllClientData':
                continue
            filePath = _path + '\\' + fileName

            with open(filePath, 'r', encoding='utf8') as f:
                data[file] = json.load(f)

        exportPath = _path + '\\' + 'AllClientData.json'
        with open(exportPath, 'w', encoding='utf8') as f:
            json.dump(data, f, ensure_ascii=False)

        warring_msg = QtWidgets.QMessageBox.information(self, "提示", "{}个客户端表成功".format(len(data.keys())),
                                                        QtWidgets.QMessageBox.Yes)

    def _exportData(self, scene):
        current_row = self.tableWidget_3.currentRow()
        # 获取文件名称
        xlsxPath = self.sheet_path[current_row]
        filename = xlsxPath.split('/')[-1].split('.')[0]
        if current_row in self.sheet_path.keys():
            wb = openpyxl.load_workbook(xlsxPath)
            ws = wb.worksheets[0]
            ws_info = wb.worksheets[1]
            if ws_info.title != '属性':
                warring_msg = QtWidgets.QMessageBox.information(self, "警告", "第二张表的title不是 属性，请检查",
                                                                QtWidgets.QMessageBox.Yes)
                return
            time_2 = time.time()
            ws_info_data = {}
            for rx in range(2, ws_info.max_row + 1):
                w1 = ws_info.cell(row=rx, column=1).value
                w2 = ws_info.cell(row=rx, column=2).value.replace(' ', '').upper()
                w3 = ws_info.cell(row=rx, column=3).value
                w4 = ws_info.cell(row=rx, column=4).value
                w5 = ws_info.cell(row=rx, column=5).value
                w6 = ws_info.cell(row=rx, column=6).value
                if w1 is not None:
                    tmp_dict = {'param': w2, 'type': w3, 'note': w4, 'info': w5, 'scene': w6}
                    ws_info_data[w1] = tmp_dict

            # 检查表格字段与属性一一对应
            for i in range(1, ws.max_column + 1):
                tmp = ws.cell(1, i).value
                if tmp and (tmp not in ws_info_data.keys()):
                    warring_msg = QtWidgets.QMessageBox.information(self, "警告", "配置表错误，字段与变量没有一一对应",
                                                                    QtWidgets.QMessageBox.Yes)
                    return
            # 输出ws_data
            ws_data = {}
            self.progressBar.setRange(1, ws.max_row - 1)
            for i in range(2, ws.max_row + 1):
                self.progressBar.setValue(i - 1)
                tmp_dict = {}
                # 空行跳过
                tmp_row = ws.cell(i, 1).value
                if tmp_row is None:
                    break
                # 处理数据
                for j in range(2, ws.max_column + 1):
                    tmp = ws.cell(1, j).value
                    # 空列数据跳出循环
                    if not tmp:
                        break
                    # 不需要的列就跳过
                    if ws_info_data[tmp]['scene'] not in scene:
                        continue
                    # 转化数据类型
                    _cellData = ws.cell(row=i, column=j).value
                    try:
                        dataType = ws_info_data[tmp]['type']
                        if dataType == 'string':
                            cellData = str(_cellData)
                        elif dataType == 'uint':
                            cellData = int(_cellData)
                        elif dataType == 'string[|]':
                            cellData = str(_cellData).split('|')
                        elif dataType == 'uint[|]':
                            cellData = [int(x) for x in str(_cellData).split('|')]
                        else:
                            warring_msg = QtWidgets.QMessageBox.information(self, "警告",
                                                                            "属性表中的变量类型只能是string|uint|sting[|]|uint[|]",
                                                                            QtWidgets.QMessageBox.Yes)
                            return
                    except:
                        warring_msg = QtWidgets.QMessageBox.information(self, "警告",
                                                                        "属性表中的变量类型错误或者有效数据区域存在空值",
                                                                        QtWidgets.QMessageBox.Yes)
                        return

                    tmp_dict[ws_info_data[tmp]['param']] = cellData
                # 导出数据
                ws_data[str(ws.cell(i, 1).value).upper()] = tmp_dict
            # 把不需要的字段删除掉
            _delete_key = []
            for i in ws_info_data.keys():
                if ws_info_data[i]['scene'] not in scene:
                    _delete_key.append(i)
            for i in _delete_key:
                del ws_info_data[i]
            return ws_data, filename, ws_info_data
        else:
            warring_msg = QtWidgets.QMessageBox.information(self, "警告", "文件未选中，请检查文件状态", QtWidgets.QMessageBox.Yes)
            return

    def exportClientData(self):
        param = ('client', 'all')
        data = self._exportData(param)
        if data is None:
            return
        ws_data, filename, ws_info_data = data
        # 字段分组
        param_group = [[], [], []]
        param_count = []
        param_weight = []

        for i in ws_info_data.keys():
            _param = ws_info_data[i]['param']
            if '_ID' in _param:
                param_group[0].append(_param.replace('_ID', ''))
            elif '_COUNT' in _param:
                param_group[1].append(_param.replace('_COUNT', ''))
            elif '_WEIGHT' in _param:
                param_group[2].append(_param.replace('_WEIGHT', ''))
            else:
                pass

        for i in param_group[0]:
            if i in param_group[1]:
                param_count.append((i + '_ID', i + '_COUNT'))
            if i in param_group[2]:
                param_weight.append((i + '_ID', i + '_WEIGHT'))

        for i in ws_data.keys():
            for j in param_count:
                _id = ws_data[i][j[0]]
                _count = ws_data[i][j[1]]
                _data = {}
                for k in range(len(_id)):
                    _data[_id[k]] = _count[k]
                ws_data[i][j[1]] = _data
            for j in param_weight:
                _id = ws_data[i][j[0]]
                _weight = ws_data[i][j[1]]
                _data = {}
                for k in range(len(_id)):
                    _data[_id[k]] = _weight[k]
                ws_data[i][j[1]] = _data
            for j in param_count:
                del ws_data[i][j[0]]
            for j in param_weight:
                tmp = j[0]
                if tmp in ws_data[i].keys():
                    del ws_data[i][tmp]
            # 输出文件
        _path = os.getcwd() + '\\' + 'client'
        if not os.path.exists(_path):
            os.makedirs(_path)
        filepath = _path + '\\' + filename + '.json'
        with open(filepath, 'w', encoding="utf-8") as f:
            # f.write('DATA=')
            json.dump(ws_data, f, ensure_ascii=False)
        warring_msg = QtWidgets.QMessageBox.information(self, "提示", "导出客户端表成功", QtWidgets.QMessageBox.Yes)

    def exportServerData(self):
        param = ('server', 'all')
        data = self._exportData(param)
        if data is None:
            return
        ws_data, filename, ws_info_data = data
        # 字段分组
        param_group = [[], [], []]
        param_count = []
        param_weight = []
        for i in ws_info_data.keys():
            _param = ws_info_data[i]['param']
            if '_ID' in _param:
                param_group[0].append(_param.replace('_ID', ''))
            elif '_COUNT' in _param:
                param_group[1].append(_param.replace('_COUNT', ''))
            elif '_WEIGHT' in _param:
                param_group[2].append(_param.replace('_WEIGHT', ''))
            else:
                pass
        for i in param_group[0]:
            if i in param_group[1]:
                param_count.append((i + '_ID', i + '_COUNT'))
            if i in param_group[2]:
                param_weight.append((i + '_ID', i + '_WEIGHT'))
        for i in ws_data.keys():
            for j in param_count:
                _id = ws_data[i][j[0]]
                _count = ws_data[i][j[1]]
                _data = {}
                for k in range(len(_id)):
                    _data[_id[k]] = _count[k]
                ws_data[i][j[1]] = _data
            for j in param_weight:
                _id = ws_data[i][j[0]]
                _weight = ws_data[i][j[1]]
                _data = {}
                for k in range(len(_id)):
                    _data[_id[k]] = _weight[k]
                ws_data[i][j[1]] = _data
            for j in param_count:
                del ws_data[i][j[0]]
            for j in param_weight:
                tmp = j[0]
                if tmp in ws_data[i].keys():
                    del ws_data[i][tmp]
        _path = os.getcwd() + '\\' + 'server'
        if not os.path.exists(_path):
            os.makedirs(_path)
        filepath = _path + '\\' + filename + '.py'
        with open(filepath, 'w', encoding="utf-8") as f:
            f.write('DATA=')
            json.dump(ws_data, f, ensure_ascii=False)
        warring_msg = QtWidgets.QMessageBox.information(self, "提示", "导出服务器表成功", QtWidgets.QMessageBox.Yes)

    def resert(self):
        self.item_file_dir = None
        self.pushButton.setEnabled(True)
        self.label_9.setText("请输入$item.xlsx文件的路径")
        self.tableWidget.clear()
        self.tableWidget_2.clear()
        self.tableWidget_3.clear()
        self.tableWidget_4.clear()
        self.sheet_path = {}
        self.pushButton_2.setDisabled(True)
        self.pushButton_4.setDisabled(True)
        self.pushButton_6.setDisabled(True)
        self.pushButton_7.setDisabled(True)
        self.pushButton_12.setDisabled(True)
        self.pushButton_13.setDisabled(True)
        self.pushButton_14.setDisabled(True)
        self.pushButton_15.setDisabled(True)
