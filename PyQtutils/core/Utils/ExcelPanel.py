"""
@author pipixia
@ info
将EXCEL整张表数据复制到QtableWidget上
"""
from openpyxl.worksheet import Worksheet
from PyQt5 import QtWidgets, QtCore
import math, threading
from core.Utils.TimerMake import TimerMake


def _cycle_letter(arr, level):
    """
    :param arr: 需要循环加字母的数组
    :param level: 需要加的层级
    :return:
    """
    tempArr = []
    letterArr = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'G', 'K', 'L', 'M', 'N', 'O', 'P', \
                 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    arrNum = len(arr)
    if (level == 0 or arrNum == 0):
        return letterArr
    for index in range(arrNum):
        for letter in letterArr:
            tempArr.append(arr[index] + letter)
    return tempArr


def _reduce_excel_col_name(num):
    """
    :param num: 需要生成的Excel列名称数目
    :return:['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'G', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',~~~~]
    """
    tempVal = 1
    level = 1
    while (tempVal):
        tempVal = num / (math.pow(26, level))
        if (tempVal > 1):
            level += 1
        else:
            break

    excelArr = []
    tempArr = []
    for index in range(level):
        tempArr = _cycle_letter(tempArr, index)
        for numIndex in range(len(tempArr)):
            if (len(excelArr) < num):
                excelArr.append(tempArr[numIndex])
            else:
                return excelArr
    return excelArr


class WorkThread(threading.Thread):
    def __init__(self, minrow, maxrow, ws, QTableWidgets, lock):
        super().__init__()
        self.minrow = minrow
        self.maxrow = maxrow
        self.ws = ws
        self.QTableWidgets = QTableWidgets
        self.lock = lock

    def run(self):
        for i in range(self.minrow, self.maxrow):
            self.lock.acquire()
            global jindu
            jindu += 1
            self.lock.release()
            for j in range(self.QTableWidgets.columnCount()):
                self.QTableWidgets.setItem(i, j,
                                           QtWidgets.QTableWidgetItem(str(self.ws.cell(row=i + 1, column=j + 1).value)))


def ExceltoQTableWidgets(ws, QTableWidgets, bool=False, selectmode=0, selectnum=0, progressBar=None):
    """
    by pipixia
    :param ws: EXCEL表对象
    :param QTableWidgets: 数据显示的对象
    :param bool:表格是否可以编辑，True-可编辑
    :param selectmode:表格选中方式，0 默认；1-整行选中；2-整列选中
    :param selectnum:单选模式 ，0默认，1-只能选择一个对象,2-正常情况只能选择1个对象，但按下Ctrl或Shift键后，可以跳行多选
    :param progressBar:进度条对象，默认没有，反应处理速度
    :return:
    """
    if isinstance(ws, Worksheet) and isinstance(QTableWidgets, QtWidgets.QTableWidget):
        QTableWidgets.setRowCount(ws.max_row)
        QTableWidgets.setColumnCount(ws.max_column)
        QTableWidgets.setVerticalHeaderLabels([str(x) for x in range(1, ws.max_row + 1)])
        QTableWidgets.setHorizontalHeaderLabels(_reduce_excel_col_name(ws.max_column))
        # ------------------选中模式------------------------------
        if selectmode == 1:
            QTableWidgets.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)  # 整行选中
        elif selectmode == 2:
            QTableWidgets.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectColumns)  # 整列选中
        else:
            pass
        # --------------------选择模式------------------------------
        if selectnum == 1:
            QTableWidgets.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)  # 单选模式
        elif selectnum == 2:
            # 正常情况下是单选，但按下Ctrl或Shift键后，可以跳行多选
            QTableWidgets.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        else:
            pass
        # ------------------编辑模式-------------------------------
        if not bool:
            QTableWidgets.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)  # 设置表格不可编辑
        # -------------------进度条判断----------------------------
        if not isinstance(progressBar, QtWidgets.QProgressBar):  # 判断是否需要显示进度条
            progressBar = QtWidgets.QProgressBar()  # 避免语法错误
        progressBar.setRange(0, ws.max_row - 1)
        # ----------判断需要加载的数据量，超过500行，5个线程同时加载-----
        if ws.max_row > 500:
            tmp = ws.max_row // 5
            rowlist = [[0, tmp], [tmp, 2 * tmp], [2 * tmp, 3 * tmp], [3 * tmp, 4 * tmp],
                       [4 * tmp, ws.max_row]]
            tlist = []
            lock = threading.Lock()
            global jindu
            jindu = -1
            for i in range(5):
                t = WorkThread(rowlist[i][0], rowlist[i][1], ws=ws, QTableWidgets=QTableWidgets, lock=lock)
                tlist.append(t)
            for i in tlist:
                i.start()
            for i in tlist:
                i.join()
            progressBar.setValue(jindu)
        else:  # 加载量小于500行，主线程加载即可
            for i in range(ws.max_row):
                progressBar.setValue(i)
                for j in range(ws.max_column):
                    QTableWidgets.setItem(i, j,
                                          QtWidgets.QTableWidgetItem(str(ws.cell(row=i + 1, column=j + 1).value)))
    else:
        raise TypeError('类型错误,请检查传入的对象类型')
